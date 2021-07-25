from pathlib import Path
import openpyxl

fileNameData = 'Data.xlsx'  # tên file (.xlsx) lưu thông tin sản phẩm
fileNameOrder = 'OrderList.xlsx'  # tên file (.xlsx) lưu thông đơn đặt hàng
pathR = ''  # đường dẫn đến file (.xlsx) lưu thông tin sản phẩm
pathW = ''  # đường dẫn đến file (.xlsx) lưu thông đơn đặt hàng


def read_file():
    fileNameFullPath = Path(pathR, fileNameData)
    try:
        wb_obj = openpyxl.load_workbook(fileNameFullPath)
        sheet = wb_obj.active
        print('# --------- Read data successfully! ---------')
    except Exception as e:
        print('# --------- Read data failed! ---------')
        print('!!!Error: ', type(e), e)
        print(e.args)
    finally:
        return sheet


def save_order(order_customer_name, order_phone_number,
               ordered_product_name,
               amount_order, total_order_amount,
               order_date, order_address):
    fileNameFullPath = Path(pathW, fileNameOrder)
    wb_obj = openpyxl.load_workbook(fileNameFullPath)
    sheet = wb_obj.active
    # Thêm 1 hàng giá trị vào file
    sheet.append(
        [order_customer_name, order_phone_number,
         ordered_product_name,
         amount_order, total_order_amount, order_date,
         order_address])
    # Lưu file
    try:
        wb_obj.save(fileNameFullPath)
        print('# --------- Order saved successfully! ---------')
    except Exception as e:
        print('# --------- Order saved failed! ---------')
        print('!!!Error: ', type(e), e)
        print(e.args)
        return False
    finally:
        return True


def get_familiar_customers(order_phone_number):
    fileNameFullPath = Path(pathW, fileNameOrder)
    wb_obj = openpyxl.load_workbook(fileNameFullPath)
    sheet = wb_obj.active
    order_customer_name = None
    order_address = None
    for i in sheet.iter_rows(max_row=sheet.max_row):
        if order_phone_number in i[1].value:
            order_customer_name = i[0].value
            order_address = i[6].value
    infoCustomer = [order_customer_name, order_address]
    if order_phone_number is None:
        print('Khách hàng mới: ', order_phone_number)
    else:
        print('Thông tin khách đã từng mua: ', infoCustomer, ', ', order_phone_number)
    return infoCustomer
