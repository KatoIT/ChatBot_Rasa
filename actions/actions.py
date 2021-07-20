from rasa_sdk import FormValidationAction
from rasa_sdk.events import EventType
import requests
import openpyxl
import datetime
from pathlib import Path
from rasa_sdk.types import DomainDict
from typing import Any, Text, Dict, List
from rasa_sdk import Action, Tracker
from rasa_sdk.executor import CollectingDispatcher
from rasa_sdk.events import SlotSet, FollowupAction, AllSlotsReset
import phonenumbers


# ----------------- Answer Request Counselling ------------------------#
class ActionAnswer(Action):

    def name(self) -> Text:
        return "action_answer"

    def run(self,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        # mở file
        xlsx_file = Path('', 'Data.xlsx')  # file
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active
        if tracker.get_slot("product_name_en") is None:
            # nếu chưa nhận được tên sản phẩm sẽ yêu cầu khách hàng chọn sản phẩm cần tư vấn.
            dispatcher.utter_message(
                response="utter_ask_product_name_en"
            )
        elif tracker.get_slot("request_counselling") is None:
            # Nếu chưa nhận được yêu cầu tư vấn từ khách sẽ hỏi khách cần tư vấn gì?
            dispatcher.utter_message(
                response="utter_ask_request_counselling"
            )
        else:
            # Xử lý yêu cầu của khách
            requestCustom = tracker.get_slot("request_counselling")
            productName = tracker.get_slot("product_name_en")
            content = "Rất xin lỗi!!!\nShop hiện chưa có thông tin về vấn đề này."
            if "bánh" in str(productName).lower():
                # Tư vấn theo sản phẩm Bánh dinh dưỡng Hebi
                if sheet["B" + requestCustom].value is not None:
                    # Nếu vị trí đó không có thông tin sẽ trả về content ban đầu.
                    content = str(sheet["B" + requestCustom].value)
            elif "men" in str(productName).lower():
                # Tư vấn theo sản phẩm Men vi khuẩn sống Việt Nhật
                if sheet["C" + requestCustom].value is not None:
                    content = str(sheet["C" + requestCustom].value)
            # Trả về thông tin khách cần
            dispatcher.utter_message(text=content)
            return [SlotSet("request_counselling", None)]
        return []


# ----------------- Set slot value ------------------------#
class ActionComfirm(Action):
    def name(self) -> Text:
        return "action_confirm_order"

    def run(self,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        #
        xlsx_file = Path('', 'OrderList.xlsx')  # file
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active
        # lấy thông tin đơn hàng từ slot
        name = str(tracker.get_slot("name"))
        phone = str(tracker.get_slot("phone"))
        product_name = str(tracker.get_slot("product_name_en"))
        amount = str(tracker.get_slot("amount"))
        total = str(tracker.get_slot("total"))
        date = str(datetime.datetime.today())
        address = str(tracker.get_slot("ward")) + ', ' + str(tracker.get_slot("district")) + ', ' + str(tracker.get_slot("province"))
        # Thêm 1 hàng giá trị vào file
        sheet.append([name, phone, product_name, amount, total, date, address])
        # Lưu file
        wb_obj.save('OrderList.xlsx')
        # Hiển thị xác nhận đã lưu đơn
        dispatcher.utter_message(
            text="Đơn hàng của bạn đã được lưu lại✅\nShop sẽ liên hệ xác nhận trong vòng 24h, vui lòng chú ý điện thoại của bạn.\n")
        return []


# ----------------- Order Form ------------------------#

class ActionOrderForm(Action):

    def name(self) -> Text:
        return "action_order_form"

    def run(
            self,
            dispatcher: "CollectingDispatcher",
            tracker: Tracker,
            domain: "DomainDict",
    ) -> List[Dict[Text, Any]]:
        required_slots = ["name", "phone", "product_name_en", "amount", "province", "district", "ward"]
        for slot_name in required_slots:
            if tracker.get_slot(slot_name) is None:
                # nếu như có một thành phần nào đó rỗng thì sẽ phải điền đủ
                return [SlotSet("requested_slot", slot_name)]
        # nếu đã có giá trị mà hàm này đk gọi đến thì nó set slot này bằng None
        return [SlotSet("requested_slot", None)]
        # dispatcher.utter_message(text= tracker.get_slot("phone_number"))


# ----------------- Validate From ------------------------#
class ValidateActionOrderForm(FormValidationAction):
    id_province = ""
    id_district = ""
    header = {'token': '82a0da54-c84b-11eb-bb70-b6be8148d819', 'Content-Type': 'application/json'}

    def name(self) -> Text:
        return "validate_action_order_form"

    def province_db(self) -> List[dict]:
        r = requests.get(url='https://online-gateway.ghn.vn/shiip/public-api/master-data/province', headers=self.header)
        data = r.json()['data']
        list_province = []
        for province in data:
            try:
                prodict = {
                    "id_province": province['ProvinceID'],
                    "NameExtension": province['NameExtension']
                }
                list_province.append(dict(prodict))
            except:
                print("Có Tỉnh/Thành Đặc Biệt")
        return list_province

    def district_db(self) -> List[dict]:
        list_district = []
        if self.id_province:
            param = {'province_id': str(self.id_province)}
            r = requests.get(url='https://online-gateway.ghn.vn/shiip/public-api/master-data/district',
                             headers=self.header,
                             params=param)
            data = r.json()['data']
            for district in data:
                try:
                    disdic = {
                        "id_district": district['DistrictID'],
                        "NameExtension": district['NameExtension']
                    }
                    list_district.append(dict(disdic))
                except:
                    print("Có Quận/Huyện Đặc Biệt")
        return list_district

    def ward_db(self) -> List[dict]:
        list_ward = []
        if self.id_district:
            param = {'district_id': str(self.id_district)}
            r = requests.get(url='https://online-gateway.ghn.vn/shiip/public-api/master-data/ward',
                             headers=self.header,
                             params=param)
            data = r.json()['data']
            for ward in data:
                try:
                    warddict = {
                        "NameExtension": ward['NameExtension']
                    }
                    list_ward.append(dict(warddict))
                except:
                    print("Có Xã/Phường Đặc Biệt")
        return list_ward

    def validate_province(
            self,
            slot_value: Any,
            dispatcher: "CollectingDispatcher",
            tracker: "Tracker",
            domain: "DomainDict",
    ) -> List[EventType]:
        for province in self.province_db():
            for namePro in province.get('NameExtension'):
                if slot_value[0].lower() in str(namePro).lower():
                    print('Tỉnh ', slot_value)
                    self.id_province = province['id_province']
                    return {"province": slot_value[0], "district": None}
                else:
                    continue
        return {"province": None, "district": None}

    def validate_district(
            self,
            slot_value: Any,
            dispatcher: "CollectingDispatcher",
            tracker: "Tracker",
            domain: "DomainDict",
    ) -> List[EventType]:
        for district in self.district_db():
            for nameDis in district['NameExtension']:
                if slot_value[0].lower() in str(nameDis).lower():
                    print('Huyện ', slot_value)
                    self.id_district = district['id_district']
                    return {"district": slot_value[0]}
                else:
                    continue
        return {"district": None}

    def validate_ward(
            self,
            slot_value: Any,
            dispatcher: "CollectingDispatcher",
            tracker: "Tracker",
            domain: "DomainDict",
    ) -> List[EventType]:
        for ward in self.ward_db():
            for nameWard in ward['NameExtension']:
                if slot_value[0].lower() in str(nameWard).lower():
                    print('Xã ', slot_value)
                    return {"ward": slot_value}
                else:
                    continue
        return {"ward": None}
# Kiểm tra số điện thoại #
    def validate_phone(
            self,
            slot_value: Any,
            dispatcher: "CollectingDispatcher",
            tracker: "Tracker",
            domain: "DomainDict",
    ) -> List[EventType]:
        countSDT = False
        for match in slot_value.PhoneNumberMatcher(slot_value, "VN"):
            countSDT = True
        if countSDT:
            return {"phone": slot_value[1]}
        return {"phone": None}
# Kiểm tra số lượng đặt hàng #
    def validate_amount(
            self,
            slot_value: Any,
            dispatcher: "CollectingDispatcher",
            tracker: "Tracker",
            domain: "DomainDict",
    ) -> List[EventType]:
        if int(slot_value[1] > 0):
            return {"amount": slot_value[1]}
        return {"amount": None}


# ----------------- Submit Form ------------------------#

class ActionSubmit(Action):

    def name(self) -> Text:
        return "action_submit"

    def run(
            self,
            dispatcher: "CollectingDispatcher",
            tracker: Tracker,
            domain: "DomainDict",
    ) -> List[Dict[Text, Any]]:
        xlsx_file = Path('', 'Data.xlsx')  # file
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active
        if "bánh" in str(tracker.get_slot("productName")).lower():
            a = str(int(tracker.get_slot("amount")) * int(sheet["D10"]))
        elif "men" in str(tracker.get_slot("productName")).lower():
            a = str(int(tracker.get_slot("amount")) * int(sheet["E10"]))
        dispatcher.utter_message(response="utter_order_details",
                                 name=tracker.get_slot("name"),
                                 phone=tracker.get_slot("phone"),
                                 amount=tracker.get_slot("amount"),
                                 product_name_en=tracker.get_slot("product_name_en"),
                                 total=a)
        return [SlotSet('total', a)]


# ----------------- Reset slot ------------------------#

class ActionResetSlot(Action):
    def name(self) -> Text:
        return "action_reset_slot"

    def run(
            self,
            dispatcher: "CollectingDispatcher",
            tracker: Tracker,
            domain: "DomainDict",
    ) -> List[Dict[Text, Any]]:
        dispatcher.utter_message(text="tất cả thông tin đã được xóa rồi nhé bạn!")
        return [AllSlotsReset()]


# ----------------- 1 Set value slot composition  ------------------------#
class ActionSetSlotComposition(Action):

    def name(self) -> Text:
        return "action_set_slot_composition"

    async def run(self, dispatcher, tracker: Tracker, domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        # custom behavior
        return [SlotSet("request_counselling", "2")]


# ----------------- 2  Set value slot ------------------------#
class ActionSetSlotTaste(Action):

    def name(self) -> Text:
        return "action_set_slot_taste"

    async def run(self, dispatcher, tracker: Tracker, domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        # custom behavior
        return [SlotSet("request_counselling", "3")]


# ----------------- 3  Set value slot ------------------------#
class ActionSetSlotEffects(Action):

    def name(self) -> Text:
        return "action_set_slot_effects"

    async def run(self, dispatcher, tracker: Tracker, domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        # custom behavior
        return [SlotSet("request_counselling", "4")]


# ----------------- 4  Set value slot ------------------------#
class ActionSetSlotContraindications(Action):

    def name(self) -> Text:
        return "action_set_slot_contraindications"

    async def run(self, dispatcher, tracker: Tracker, domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        # custom behavior
        return [SlotSet("request_counselling", "5")]


# ----------------- 5  Set value slot ------------------------#
class ActionSetSlotUserManual(Action):

    def name(self) -> Text:
        return "action_set_slot_user_manual"

    async def run(self, dispatcher, tracker: Tracker, domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        # custom behavior
        if tracker.get_slot("age") is None:
            return [SlotSet("request_counselling", "6")]
        else:
            age = int(tracker.get_slot("age"))
            if age < 2:
                return [SlotSet("request_counselling", "16")]
            elif age > 15:
                return [SlotSet("request_counselling", "18")]
            else:
                return [SlotSet("request_counselling", "17")]


# ----------------- 6  Set value slot ------------------------#
class ActionSetSlotStorage(Action):

    def name(self) -> Text:
        return "action_set_slot_storage"

    async def run(self, dispatcher, tracker: Tracker, domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        # custom behavior
        return [SlotSet("request_counselling", "7")]


# ----------------- 7  Set value slot ------------------------#
class ActionSetSlotMadeIn(Action):

    def name(self) -> Text:
        return "action_set_slot_made_in"

    async def run(self, dispatcher, tracker: Tracker, domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        # custom behavior
        return [SlotSet("request_counselling", "8")]


# ----------------- 8  Set value slot ------------------------#
class ActionSetSlotSale(Action):

    def name(self) -> Text:
        return "action_set_slot_sale"

    async def run(self, dispatcher, tracker: Tracker, domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        # custom behavior
        return [SlotSet("request_counselling", "9")]


# ----------------- 9  Set value slot ------------------------#
class ActionSetSlotPrice(Action):

    def name(self) -> Text:
        return "action_set_slot_price"

    async def run(self, dispatcher, tracker: Tracker, domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        # custom behavior
        return [SlotSet("request_counselling", "10")]


# ----------------- 10  Set value slot ------------------------#
class ActionSetSlotNominations(Action):

    def name(self) -> Text:
        return "action_set_slot_nominations"

    async def run(self, dispatcher, tracker: Tracker, domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        # custom behavior
        return [SlotSet("request_counselling", "11")]


# ----------------- 11  Set value slot ------------------------#
class ActionSetSlotRecognizingSigns(Action):

    def name(self) -> Text:
        return "action_set_slot_recognizing_signs"

    async def run(self, dispatcher, tracker: Tracker, domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        # custom behavior
        return [SlotSet("request_counselling", "12")]


# ----------------- 12  Set value slot ------------------------#
class ActionSetSlotShip(Action):

    def name(self) -> Text:
        return "action_set_slot_ship"

    async def run(self, dispatcher, tracker: Tracker, domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        # custom behavior
        return [SlotSet("request_counselling", "13")]


# ----------------- 13  Set value slot ------------------------#
class ActionSetSlotUserObject(Action):

    def name(self) -> Text:
        return "action_set_slot_user_object"

    async def run(self, dispatcher, tracker: Tracker, domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        # custom behavior
        return [SlotSet("request_counselling", "14")]
